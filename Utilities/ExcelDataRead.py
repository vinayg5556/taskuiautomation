import openpyxl
from Utilities.configReader import read_config


def get_row_data(file_path, sheet_name):
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.max_row


def get_column_data(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.max_column


def read_data(file_path, sheet_name, row_num, col_num):
    print(file_path)
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    return ws.cell(row=row_num, column=col_num).value


def write_data(file_path, sheet_name, row_num, col_num, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    ws.cell(row=row_num, column=col_num).value = data
    wb.save(file_path)


def get_input_data(sheet_name):
    path = read_config("Input", "path")
    rows = get_row_data(path, sheet_name)
    col = get_column_data(path, sheet_name)
    mainList = []
    for i in range(2, rows + 1):
        dataList = []
        for j in range(1, col + 1):
            data = read_data(path, sheet_name, i, j)
            print(data)
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    print(mainList)
    return mainList
